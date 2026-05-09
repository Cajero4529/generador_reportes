import os from flask import Flask, render_template, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
import io
import os

app = Flask(__name__)
CORS(app)

class GeneradorReportes:
    def __init__(self):
        self.naranja = 'FF8C00'
        self.bordó = '800020'
        self.gris_claro = 'E8E8E8'
        self.gris_oscuro = 'D3D3D3'
    
    def procesar(self, archivo_bytes, pto_venta, z_codigo):
        """Procesa el archivo y genera el reporte"""
        try:
            df = pd.read_excel(io.BytesIO(archivo_bytes), sheet_name='Operaciones')
            df['N° Cupón'] = df['N° Cupón'].astype(str)
            
            tickets_con_articulos = df.groupby('N° Caja').agg({
                'Código': lambda x: list(x),
                'Título': lambda x: list(x),
                'Total': 'first',
                'N° Cupón': 'first'
            }).reset_index()
            
            tickets_con_articulos['N_Articulos'] = tickets_con_articulos['Código'].apply(len)
            
            tickets_1art = tickets_con_articulos[tickets_con_articulos['N_Articulos'] == 1].sort_values('Total', ascending=False).reset_index(drop=True)
            tickets_2art = tickets_con_articulos[tickets_con_articulos['N_Articulos'] == 2].sort_values('Total', ascending=False).reset_index(drop=True)
            
            lotes = self._crear_lotes(tickets_1art, tickets_2art)
            self._procesar_articulos(lotes, df, tickets_con_articulos)
            
            excel_bytes = self._generar_excel(lotes, pto_venta, z_codigo)
            return excel_bytes
            
        except Exception as e:
            print(f"Error: {str(e)}")
            raise
    
    def _crear_lotes(self, tickets_1art, tickets_2art):
        """Crea lotes respetando distribución Benford"""
        lotes = []
        lote_actual = {'numero': 1, 'cajas': [], 'items': 0, 'monto': 0, 'composicion': {'1art': 0, '2art': 0}}
        
        idx_1art = 0
        idx_2art = 0
        total_tickets = len(tickets_1art) + len(tickets_2art)
        tickets_procesados = 0
        
        while tickets_procesados < total_tickets:
            items_faltantes = 9 - lote_actual['items']
            
            if items_faltantes == 0:
                lotes.append(lote_actual)
                lote_actual = {'numero': len(lotes) + 1, 'cajas': [], 'items': 0, 'monto': 0, 'composicion': {'1art': 0, '2art': 0}}
                items_faltantes = 9
            
            proporcion_1art_actual = (lote_actual['composicion']['1art'] / 9) if lote_actual['items'] > 0 else 0.73
            added = False
            
            if idx_1art < len(tickets_1art) and items_faltantes >= 1 and proporcion_1art_actual < 0.73:
                ticket = tickets_1art.iloc[idx_1art]
                lote_actual['cajas'].append(ticket['N° Caja'])
                lote_actual['items'] += 1
                lote_actual['monto'] += ticket['Total']
                lote_actual['composicion']['1art'] += 1
                idx_1art += 1
                tickets_procesados += 1
                added = True
            elif idx_2art < len(tickets_2art) and items_faltantes >= 2:
                ticket = tickets_2art.iloc[idx_2art]
                lote_actual['cajas'].append(ticket['N° Caja'])
                lote_actual['items'] += 2
                lote_actual['monto'] += ticket['Total']
                lote_actual['composicion']['2art'] += 1
                idx_2art += 1
                tickets_procesados += 1
                added = True
            elif idx_1art < len(tickets_1art) and items_faltantes >= 1:
                ticket = tickets_1art.iloc[idx_1art]
                lote_actual['cajas'].append(ticket['N° Caja'])
                lote_actual['items'] += 1
                lote_actual['monto'] += ticket['Total']
                lote_actual['composicion']['1art'] += 1
                idx_1art += 1
                tickets_procesados += 1
                added = True
            
            if not added and lote_actual['cajas']:
                lotes.append(lote_actual)
                lote_actual = {'numero': len(lotes) + 1, 'cajas': [], 'items': 0, 'monto': 0, 'composicion': {'1art': 0, '2art': 0}}
        
        if lote_actual['cajas']:
            lotes.append(lote_actual)
        
        return lotes
    
    def _procesar_articulos(self, lotes, df, tickets_con_articulos):
        """Procesa artículos para cada lote"""
        for lote in lotes:
            lote['cajas'] = sorted(lote['cajas'], reverse=True)
            lote['cajas_cupones_articulos'] = []
            
            for caja_id in lote['cajas']:
                caja_data = tickets_con_articulos[tickets_con_articulos['N° Caja'] == caja_id].iloc[0]
                cupón = str(caja_data['N° Cupón'])
                
                articulos_caja = []
                caja_df = df[df['N° Caja'] == caja_id]
                for _, row in caja_df.iterrows():
                    articulos_caja.append({
                        'Código': row['Código'],
                        'Título': row['Título'],
                        'Cantidad': row['Cantidad'],
                        'Precio Unit.': row['Precio Unit.'],
                        'Subtotal': row['Subtotal'],
                        'Cupón': cupón
                    })
                
                lote['cajas_cupones_articulos'].append({
                    'caja': caja_id,
                    'cupón': cupón,
                    'articulos': articulos_caja
                })
            
            articulos_lote = []
            for caja_info in lote['cajas_cupones_articulos']:
                articulos_lote.extend(caja_info['articulos'])
            
            df_lote = pd.DataFrame(articulos_lote)
            lote['articulos_agrupados'] = df_lote.groupby(['Código', 'Título']).agg({
                'Cantidad': 'sum',
                'Precio Unit.': 'first',
                'Subtotal': 'sum',
                'Cupón': lambda x: ', '.join(sorted(set(str(c) for c in x)))
            }).reset_index().sort_values('Cantidad', ascending=False)
    
    def _generar_excel(self, lotes, pto_venta, z_codigo):
        """Genera el archivo Excel y lo retorna como bytes"""
        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        header_fill = PatternFill(start_color=self.naranja, end_color=self.naranja, fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        lote_header_fill = PatternFill(start_color=self.gris_oscuro, end_color=self.gris_oscuro, fill_type='solid')
        lote_header_font = Font(bold=True, size=11)
        subheader_fill = PatternFill(start_color=self.gris_claro, end_color=self.gris_claro, fill_type='solid')
        subheader_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color=self.bordó, end_color=self.bordó, fill_type='solid')
        total_font = Font(bold=True, color='FFFFFF', size=11)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # HOJA 1: RESUMEN LOTES
        sheet_resumen = wb.create_sheet("Resumen Lotes", 0)
        row = 1
        
        for lote in lotes:
            sheet_resumen.merge_cells(f'A{row}:D{row}')
            cell = sheet_resumen.cell(row=row, column=1)
            cell.value = f"LOTE {lote['numero']} - Cajas: {', '.join(lote['cajas'])}"
            cell.font = lote_header_font
            cell.fill = lote_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            sheet_resumen.row_dimensions[row].height = 18
            row += 1
            
            headers = ['Nro.', 'N° Cupón', 'Detalle del Concepto', 'Abona']
            for col, header in enumerate(headers, 1):
                cell = sheet_resumen.cell(row=row, column=col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1
            
            nro_orden = 1
            for _, art in lote['articulos_agrupados'].iterrows():
                cell = sheet_resumen.cell(row=row, column=1)
                cell.value = nro_orden
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                cell = sheet_resumen.cell(row=row, column=2)
                cell.value = art['Cupón']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                cell.number_format = '@'
                
                codigo = art['Código']
                cantidad = int(art['Cantidad'])
                detalle = f"{pto_venta}/{z_codigo} / {codigo} / {cantidad}"
                
                cell = sheet_resumen.cell(row=row, column=3)
                cell.value = detalle
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
                
                cell = sheet_resumen.cell(row=row, column=4)
                cell.value = int(art['Subtotal'])
                cell.number_format = '$#,##0'
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                nro_orden += 1
                row += 1
            
            sheet_resumen.merge_cells(f'A{row}:C{row}')
            cell = sheet_resumen.cell(row=row, column=1)
            cell.value = f"TOTAL LOTE {lote['numero']}"
            cell.font = Font(bold=True)
            cell.fill = lote_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            cell = sheet_resumen.cell(row=row, column=4)
            cell.value = lote['monto']
            cell.number_format = '$#,##0'
            cell.font = Font(bold=True)
            cell.fill = lote_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            row += 2
        
        sheet_resumen.merge_cells(f'A{row}:C{row}')
        cell = sheet_resumen.cell(row=row, column=1)
        cell.value = "TOTAL GENERAL"
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        cell = sheet_resumen.cell(row=row, column=4)
        cell.value = sum(l['monto'] for l in lotes)
        cell.number_format = '$#,##0'
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        sheet_resumen.column_dimensions['A'].width = 8
        sheet_resumen.column_dimensions['B'].width = 15
        sheet_resumen.column_dimensions['C'].width = 50
        sheet_resumen.column_dimensions['D'].width = 15
        
        # HOJA 2: COMPLETO
        sheet_completo = wb.create_sheet("Completo", 1)
        
        sheet_completo['A1'] = 'DETALLE DE PRODUCTOS POR LOTE - DISTRIBUCIÓN NATURAL DE BENFORD'
        sheet_completo['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        sheet_completo['A1'].fill = header_fill
        sheet_completo.merge_cells('A1:E1')
        sheet_completo['A1'].alignment = Alignment(horizontal='center', vertical='center')
        sheet_completo.row_dimensions[1].height = 25
        
        fecha = datetime.now().strftime('%d/%m/%Y %H:%M')
        sheet_completo['A2'] = f'Procesado: {fecha} | Pto. Venta: {pto_venta} | Z: {z_codigo}'
        sheet_completo['A2'].font = Font(italic=True, size=9)
        sheet_completo.merge_cells('A2:E2')
        
        row = 4
        
        for lote in lotes:
            sheet_completo.merge_cells(f'A{row}:E{row}')
            cell = sheet_completo.cell(row=row, column=1)
            cell.value = f"LOTE {lote['numero']} - Cajas: {', '.join(lote['cajas'])}"
            cell.font = lote_header_font
            cell.fill = lote_header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
            sheet_completo.row_dimensions[row].height = 18
            row += 1
            
            headers_resumen = ['Código', 'Descripción', 'Cantidad', 'Precio Unit.', 'Monto']
            for col, header in enumerate(headers_resumen, 1):
                cell = sheet_completo.cell(row=row, column=col)
                cell.value = header
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            row += 1
            
            for _, art in lote['articulos_agrupados'].iterrows():
                cell = sheet_completo.cell(row=row, column=1)
                cell.value = art['Código']
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                cell = sheet_completo.cell(row=row, column=2)
                cell.value = art['Título']
                cell.border = border
                cell.alignment = Alignment(horizontal='left', wrap_text=True)
                
                cell = sheet_completo.cell(row=row, column=3)
                cell.value = int(art['Cantidad'])
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
                
                cell = sheet_completo.cell(row=row, column=4)
                cell.value = art['Precio Unit.']
                cell.number_format = '$#,##0'
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                cell = sheet_completo.cell(row=row, column=5)
                cell.value = int(art['Subtotal'])
                cell.number_format = '$#,##0'
                cell.border = border
                cell.alignment = Alignment(horizontal='right')
                
                row += 1
            
            cell = sheet_completo.cell(row=row, column=1)
            cell.value = "SUMA TOTAL"
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            cell = sheet_completo.cell(row=row, column=2)
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell.border = border
            
            cell = sheet_completo.cell(row=row, column=3)
            cell.value = sum(int(art['Cantidad']) for _, art in lote['articulos_agrupados'].iterrows())
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
            
            cell = sheet_completo.cell(row=row, column=4)
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell.border = border
            
            cell = sheet_completo.cell(row=row, column=5)
            cell.value = sum(int(art['Subtotal']) for _, art in lote['articulos_agrupados'].iterrows())
            cell.number_format = '$#,##0'
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            cell.border = border
            cell.alignment = Alignment(horizontal='right')
            
            row += 3
        
        sheet_completo.merge_cells(f'A{row}:B{row}')
        cell = sheet_completo.cell(row=row, column=1)
        cell.value = "TOTAL GENERAL"
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        cell = sheet_completo.cell(row=row, column=3)
        cell.fill = total_fill
        cell.border = border
        
        cell = sheet_completo.cell(row=row, column=4)
        cell.fill = total_fill
        cell.border = border
        
        cell = sheet_completo.cell(row=row, column=5)
        cell.value = sum(l['monto'] for l in lotes)
        cell.number_format = '$#,##0'
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='right')
        
        sheet_completo.column_dimensions['A'].width = 14
        sheet_completo.column_dimensions['B'].width = 45
        sheet_completo.column_dimensions['C'].width = 12
        sheet_completo.column_dimensions['D'].width = 15
        sheet_completo.column_dimensions['E'].width = 15
        
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        return excel_file

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generar', methods=['POST'])
def generar():
    try:
        if 'archivo' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        archivo = request.files['archivo']
        pto_venta = request.form.get('pto_venta')
        z_codigo = request.form.get('z_codigo')
        
        if not pto_venta or not z_codigo:
            return jsonify({'error': 'Faltan datos'}), 400
        
        archivo_bytes = archivo.read()
        generador = GeneradorReportes()
        excel_bytes = generador.procesar(archivo_bytes, pto_venta, z_codigo)
        
        return send_file(
            excel_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Reporte_Lotes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    
    except Exception as e:
        print(f"Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
